from openpyxl import Workbook
from openpyxl import load_workbook

def xlsxHandler(wbPath, targetBU, listCenters, dictCenters):
	wb = load_workbook(filename = wbPath, read_only = True)
	listSheetNames = wb.sheetnames
	ws = wb[listSheetNames[0]]

	i = 0
	centerNameIndex = 0
	centerCodeIndex = 0
	centerStatusIndex = 0
	businessUnitIndex = 0
	listCenters = []
	dictCenters = {}
	for row in ws.rows:
		i += 1
		if i == 1:
			j = 0
			for cell in row:
				j += 1
				keyStr = cell.value.replace(" ", "")
				if keyStr == "AccountName":
					centerNameIndex = j
				elif keyStr == "CenterCode":
					centerCodeIndex = j
				elif keyStr == "CenterStatus":
					centerStatusIndex = j
				elif keyStr == "BusinessUnit(Text)":
					businessUnitIndex = j
		else:
			bu = ws.cell(row = i, column = businessUnitIndex).value
			centerName = ws.cell(row = i, column = centerNameIndex).value.replace(" ", "_").replace("-", "_").replace(".", "")
			centerCode = ws.cell(row = i, column = centerCodeIndex).value
			centerStatus = ws.cell(row = i, column = centerStatusIndex).value

			if ((bu == targetBU) and (centerStatus == "Active")):
				listCenters.append(centerCode)
				dictCenters[centerName] = centerCode
	return listCenters, dictCenters